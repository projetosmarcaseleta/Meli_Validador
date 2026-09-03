"""
exporter.py – Exportador Excel multi-abas para Triagem e Exclusão em Massa
"""

from __future__ import annotations

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_triage_excel(
    items: list[dict],
    selected_ids: set[str] | None = None,
    filter_category: str = "all",
) -> io.BytesIO:
    """
    Gera uma planilha Excel com abas organizadas por categoria de triagem.
    """
    wb = openpyxl.Workbook()
    # Remove a aba default inicial
    wb.remove(wb.active)

    fill_red_head = PatternFill("solid", fgColor="EF4444")
    fill_green_head = PatternFill("solid", fgColor="10B981")
    fill_yellow_head = PatternFill("solid", fgColor="F59E0B")
    fill_blue_head = PatternFill("solid", fgColor="3B82F6")
    font_head = Font(bold=True, color="FFFFFF", size=10)

    fill_red_cell = PatternFill("solid", fgColor="FEE2E2")
    fill_green_cell = PatternFill("solid", fgColor="D1FAE5")
    fill_yellow_cell = PatternFill("solid", fgColor="FEF3C7")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    headers = [
        "SELECIONADO",
        "SKU",
        "MLB_CATALOGO",
        "MLB_TRADICIONAL",
        "CATEGORIA",
        "MOTIVO_DESCARTE_RAPIDO",
        "COR_CATALOGO",
        "COR_TRADICIONAL",
        "VOLTAGEM_CATALOGO",
        "VOLTAGEM_TRADICIONAL",
        "MODELO_CATALOGO",
        "MODELO_TRADICIONAL",
        "MARCA_CATALOGO",
        "MARCA_TRADICIONAL",
        "TAMANHO_CATALOGO",
        "TAMANHO_TRADICIONAL",
        "EAN_CATALOGO",
        "EAN_TRADICIONAL",
        "PRECO_CATALOGO",
        "PRECO_TRADICIONAL",
        "FOTOS_CAT",
        "FOTOS_TRAD",
        "TITULO_CATALOGO",
        "TITULO_TRADICIONAL",
        "LINK_CATALOGO",
        "LINK_TRADICIONAL",
    ]

    def _write_sheet(ws, sheet_items, head_fill):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = head_fill
            cell.font = font_head
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, item in enumerate(sheet_items, start=2):
            item_id = item.get("item_id", "")
            is_sel = "SIM" if (selected_ids and item_id in selected_ids) else "NÃO"
            cat = item.get("cat") or {}
            trad = item.get("trad") or {}

            row_data = [
                is_sel,
                item.get("sku", ""),
                item.get("mlb_cat", ""),
                item.get("mlb_trad", ""),
                item.get("category_label", item.get("category", "")),
                item.get("reasons_summary", ""),
                cat.get("color_raw", ""),
                trad.get("color_raw", ""),
                cat.get("voltage_raw", ""),
                trad.get("voltage_raw", ""),
                cat.get("model_raw", ""),
                trad.get("model_raw", ""),
                cat.get("brand_raw", ""),
                trad.get("brand_raw", ""),
                cat.get("size_raw", ""),
                trad.get("size_raw", ""),
                cat.get("ean", ""),
                trad.get("ean", ""),
                f"R$ {cat.get('price', 0):.2f}" if cat.get("price") else "",
                f"R$ {trad.get('price', 0):.2f}" if trad.get("price") else "",
                cat.get("image_count", 0),
                trad.get("image_count", 0),
                cat.get("title", ""),
                trad.get("title", ""),
                cat.get("permalink", ""),
                trad.get("permalink", ""),
            ]
            ws.append(row_data)

            # Estilo das células
            category = item.get("category", "")
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                c.alignment = Alignment(vertical="center")
                if col_idx in (1, 2, 3, 4, 5, 21, 22):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 6 and category == "HARD_MISMATCH":
                    c.fill = fill_red_cell
                    c.font = Font(bold=True, color="991B1B")

        # Auto fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 48)

        ws.freeze_panes = "A2"

    # Filtra por seleção se solicitado
    items_to_process = items
    if selected_ids:
        items_to_process = [it for it in items if it.get("item_id") in selected_ids]

    # Aba 1: 🔴 Descarte Imediato
    hard_items = [it for it in items_to_process if it.get("category") == "HARD_MISMATCH"]
    if hard_items or filter_category in ("all", "hard"):
        ws_hard = wb.create_sheet(title="🔴 Descarte Imediato")
        _write_sheet(ws_hard, hard_items, fill_red_head)

    # Aba 2: 🟢 Aptos para Validação
    clean_items = [it for it in items_to_process if it.get("category") == "CLEAN_MATCH"]
    if clean_items or filter_category in ("all", "clean"):
        ws_clean = wb.create_sheet(title="🟢 Aptos e Compatíveis")
        _write_sheet(ws_clean, clean_items, fill_green_head)

    # Aba 3: 🟡 Incompletos e Alertas
    inc_items = [it for it in items_to_process if it.get("category") in ("INCOMPLETE", "SOFT_DIFF")]
    if inc_items or filter_category in ("all", "incomplete"):
        ws_inc = wb.create_sheet(title="🟡 Incompletos e Alertas")
        _write_sheet(ws_inc, inc_items, fill_yellow_head)

    # Aba 4: 📊 Geral
    if filter_category == "all" and len(items_to_process) > 0:
        ws_all = wb.create_sheet(title="📊 Geral (Todos)")
        _write_sheet(ws_all, items_to_process, fill_blue_head)

    # Se nenhuma aba foi criada
    if not wb.sheetnames:
        ws_empty = wb.create_sheet(title="Vazio")
        ws_empty.append(["Nenhum anúncio encontrado para os filtros selecionados."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
