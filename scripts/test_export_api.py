"""Smoke test for /api/export (catalog mode)."""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / "ml_exporter"
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT.parent / ".env")
load_dotenv(ROOT.parent.parent / ".env")

from app import app  # noqa: E402


MOCK_ROWS = [
    [
        "SKU",
        "MLB_CATALOGO",
        "MLB_TRADICIONAL",
        "STATUS_CATALOGO",
        "STATUS_TRADICIONAL",
        "Δ_STATUS",
        "PRECO_CATALOGO",
        "PRECO_TRADICIONAL",
        "Δ_PRECO",
        "ESTOQUE_CATALOGO",
        "ESTOQUE_TRADICIONAL",
        "Δ_ESTOQUE",
        "TIPO_CATALOGO",
        "TIPO_TRADICIONAL",
        "Δ_TIPO",
        "ENVIO_CATALOGO",
        "ENVIO_TRADICIONAL",
        "Δ_ENVIO",
        "TITULO_CATALOGO",
        "TITULO_TRADICIONAL",
        "MARCA_CATALOGO",
        "MARCA_TRADICIONAL",
        "VOLTAGEM_CATALOGO",
        "VOLTAGEM_TRADICIONAL",
        "QTD_FOTOS_CATALOGO",
        "QTD_FOTOS_TRADICIONAL",
        "DIVERGENCIAS",
        "STATUS_GERAL",
        "DECISAO_AUDITORIA",
    ],
    [
        "238601800",
        "MLB5120115391",
        "MLB7479710702",
        "active",
        "active",
        "OK",
        "99.90",
        "99.90",
        "OK",
        "10",
        "10",
        "OK",
        "gold_special",
        "gold_special",
        "OK",
        "fulfillment",
        "fulfillment",
        "OK",
        "Titulo Cat",
        "Titulo Trad",
        "Marca",
        "Marca",
        "127V",
        "127V",
        "5",
        "5",
        "OK",
        "OK",
        "PENDENTE",
    ],
]


def test_export_mocked_all() -> None:
    client = app.test_client()
    payload = {
        "token": "TEST_TOKEN",
        "skus": ["238601800"],
        "mode": "catalog",
        "reviews": {"238601800": "PENDENTE"},
        "filter_decision": "all",
    }
    with patch("app.process_skus_for_catalog_excel", return_value=(MOCK_ROWS, [])):
        res = client.post("/api/export", json=payload)

    assert res.status_code == 200, res.get_json()
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert res.headers.get("X-Export-Count") == "1"

    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    ws = wb.active
    assert ws.cell(1, 1).value == "SKU"
    assert ws.cell(2, 1).value == "238601800"
    print("OK mocked export all:", len(res.data), "bytes")


def test_export_mocked_approved_empty() -> None:
    client = app.test_client()
    payload = {
        "token": "TEST_TOKEN",
        "skus": ["238601800"],
        "mode": "catalog",
        "reviews": {"238601800": "PENDENTE"},
        "filter_decision": "approved",
    }
    with patch("app.process_skus_for_catalog_excel", return_value=([MOCK_ROWS[0]], [])):
        res = client.post("/api/export", json=payload)

    assert res.status_code == 422
    print("OK approved empty -> 422")


def test_export_live() -> None:
    token = os.getenv("MELI_ACCESS_TOKEN", "").strip()
    if not token:
        print("SKIP live export: MELI_ACCESS_TOKEN not set")
        return

    client = app.test_client()
    payload = {
        "token": token,
        "skus": ["238601800"],
        "mode": "catalog",
        "reviews": {"238601800": "PENDENTE"},
        "filter_decision": "all",
    }
    res = client.post("/api/export", json=payload)
    if res.status_code != 200:
        print("FAIL live export:", res.status_code, res.get_json())
        sys.exit(1)

    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    rows = list(wb.active.iter_rows(min_row=1, max_row=2, values_only=True))
    print("OK live export:", res.headers.get("X-Export-Count"), "rows", rows[1][0] if len(rows) > 1 else "?")


def test_export_mocked_audit_items() -> None:
    client = app.test_client()
    payload = {
        "token": "TEST_TOKEN",
        "skus": ["238601800"],
        "mode": "catalog",
        "reviews": {"238601800": "PENDENTE"},
        "filter_decision": "all",
        "audit_items": [
            {
                "sku": "238601800",
                "mlb_cat": "MLB5120115391",
                "mlb_trad": "MLB7479710702",
                "ml": {"status": "active", "price": "99.90", "stock": "10", "listing_type": "gold_special", "shipping_type": "fulfillment", "title": "Cat", "brand": "Marca", "voltage": "127V", "image_count": 5},
                "any": {"status": "active", "price": "99.90", "stock": "10", "listing_type": "gold_special", "shipping_type": "fulfillment", "title": "Trad", "brand": "Marca", "voltage": "127V", "image_count": 5},
                "divergences": [],
                "status_geral": "OK",
            }
        ],
    }
    with patch("app.process_skus_for_catalog_audit") as mock_audit:
        res = client.post("/api/export", json=payload)
        mock_audit.assert_not_called()

    assert res.status_code == 200, res.get_json()
    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    assert wb.active.cell(2, 1).value == "238601800"
    print("OK export with audit_items cache (no re-fetch)")


if __name__ == "__main__":
    test_export_mocked_all()
    test_export_mocked_approved_empty()
    test_export_mocked_audit_items()
    test_export_live()
