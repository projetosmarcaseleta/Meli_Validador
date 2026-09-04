"""
app.py – Interface web Flask para Relatorios Meli + validação AnyMarket
Gera planilha Excel com campos do Mercado Livre e, opcionalmente,
colunas ANY_* / MATCH_* consultando GET /products da AnyMarket.
"""

import os
import io
import uuid
import json

from flask import Flask, render_template, request, jsonify, send_file, redirect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from anymarket_api import validate_gumga_token
from api import validate_token
from config import ANYMARKET_PLATFORM, GUMGA_TOKEN, AI_PREVALIDATION_ENABLED, MELI_TOKEN_WEBHOOK_URL, HTTP_TIMEOUT
from n8n_token import fetch_meli_token_from_n8n
from exporter import (
    process_mlbs,
    process_mlbs_for_audit,
    process_skus_for_catalog_audit,
    process_skus_for_catalog_excel,
)
from import_parser import parse_spreadsheet
from openai_vision import run_prevalidation
from prevalidation import PROMPT_VERSION, PrevalidationError, pair_from_audit_item

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "relatorios-meli-dev")
PUBLIC_EXPORT_URL = os.environ.get("PUBLIC_EXPORT_URL", "https://app.marcaseleta.shop/auditarcatalogo")


def _write_excel(rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cruzamento ML x Any"

    header_ml = PatternFill("solid", fgColor="FFE600")
    header_any = PatternFill("solid", fgColor="63B3FF")
    header_delta = PatternFill("solid", fgColor="34D399")
    header_imp = PatternFill("solid", fgColor="F472B6")
    header_dec = PatternFill("solid", fgColor="C084FC")
    header_font = Font(bold=True, color="000000")

    ok_fill = PatternFill("solid", fgColor="D1FAE5")
    err_fill = PatternFill("solid", fgColor="FECACA")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")

    headers = rows[0]
    delta_cols = {i for i, h in enumerate(headers, 1) if str(h).startswith("Δ_")}
    div_col = next((i for i, h in enumerate(headers, 1) if h == "DIVERGENCIAS"), None)
    status_col = next((i for i, h in enumerate(headers, 1) if h == "STATUS_GERAL"), None)
    dec_col = next((i for i, h in enumerate(headers, 1) if h == "DECISAO_AUDITORIA"), None)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        h = str(header)
        if h.startswith("Δ_"):
            cell.fill = header_delta
        elif h == "DECISAO_AUDITORIA":
            cell.fill = header_dec
        elif h.startswith("ANY_"):
            cell.fill = header_any
        elif h.startswith("ML_"):
            cell.fill = header_ml
        elif h.startswith("IMP_"):
            cell.fill = header_imp
        elif h in ("DIVERGENCIAS", "STATUS_GERAL"):
            cell.fill = header_delta
        else:
            cell.fill = header_ml

    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in delta_cols:
                val = str(value or "")
                if val == "OK":
                    cell.fill = ok_fill
                elif val == "DIVERGENTE":
                    cell.fill = err_fill
                elif val.startswith("AUSENTE"):
                    cell.fill = warn_fill
            if col_idx == status_col:
                val = str(value or "")
                if val == "OK":
                    cell.fill = ok_fill
                elif val == "DIVERGENTE":
                    cell.fill = err_fill
            if col_idx == dec_col:
                val = str(value or "")
                if val == "APROVADO":
                    cell.fill = ok_fill
                elif val == "REPROVADO":
                    cell.fill = err_fill
                elif val == "PENDENTE":
                    cell.fill = warn_fill
            if col_idx == div_col and str(value or "").startswith("OK"):
                cell.fill = ok_fill
            elif col_idx == div_col and value and not str(value).startswith("OK"):
                cell.fill = err_fill

    default_width = 16
    wide = {"DESCRICAO", "IMAGENS", "IMAGEM_PRINCIPAL", "DIVERGENCIAS", "LINK_ML", "TITULO", "IMP_TITULO"}
    for col_idx, header in enumerate(headers, start=1):
        label = str(header).replace("ML_", "").replace("ANY_", "").replace("Δ_", "")
        width = 55 if label in wide or header == "DIVERGENCIAS" else default_width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
@app.route("/auditarcatalogo")
@app.route("/auditarcatalogo/")
def export_index():
    return render_template(
        "index.html",
        default_gumga_token=GUMGA_TOKEN or "",
        default_any_platform=ANYMARKET_PLATFORM or "SELETA",
    )


@app.route("/api/validate_token", methods=["POST"])
@app.route("/auditarcatalogo/api/validate_token", methods=["POST"])
def api_validate_token():
    data = request.json or {}
    token = (data.get("token") or "").strip()
    if not token:
        return jsonify({"valid": False, "error": "Token vazio."}), 400

    user = validate_token(token)
    if user and user.get("id"):
        return jsonify({"valid": True, "nickname": user.get("nickname"), "id": user.get("id")})
    return jsonify({"valid": False, "error": "Token inválido ou expirado."}), 401


@app.route("/api/refresh_token", methods=["POST"])
@app.route("/auditarcatalogo/api/refresh_token", methods=["POST"])
def api_refresh_token():
    result = fetch_meli_token_from_n8n(MELI_TOKEN_WEBHOOK_URL, timeout=HTTP_TIMEOUT)
    if not result.get("ok"):
        return jsonify({"success": False, "error": result.get("error") or "Falha ao renovar o token."}), 502

    token = (result.get("token") or "").strip()
    user = validate_token(token)
    if not user or not user.get("id"):
        return jsonify({
            "success": False,
            "error": "Token renovado, mas o Mercado Livre recusou. Tente novamente.",
        }), 401

    return jsonify({
        "success": True,
        "token": token,
        "nickname": user.get("nickname"),
        "id": user.get("id"),
    })


@app.route("/api/validate_gumga", methods=["POST"])
@app.route("/auditarcatalogo/api/validate_gumga", methods=["POST"])
def api_validate_gumga():
    data = request.json or {}
    gumga = (data.get("gumga_token") or data.get("token") or "").strip()
    platform = (data.get("platform") or ANYMARKET_PLATFORM or "SELETA").strip()
    result = validate_gumga_token(gumga, platform)
    status = 200 if result.get("valid") else 401
    return jsonify(result), status


@app.route("/api/import_spreadsheet", methods=["POST"])
@app.route("/auditarcatalogo/api/import_spreadsheet", methods=["POST"])
def api_import_spreadsheet():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"success": False, "error": "Nenhum arquivo enviado."}), 400

    content = upload.read()
    if not content:
        return jsonify({"success": False, "error": "Arquivo vazio."}), 400

    rows, warnings = parse_spreadsheet(upload.filename, content)
    if not rows:
        return jsonify({
            "success": False,
            "error": warnings[0] if warnings else "Não foi possível ler a planilha.",
            "warnings": warnings,
        }), 422

    return jsonify({
        "success": True,
        "count": len(rows),
        "mlbs": [r.mlb for r in rows],
        "rows": [r.as_dict() for r in rows],
        "warnings": warnings,
    })


@app.route("/api/export", methods=["POST"])
@app.route("/auditarcatalogo/api/export", methods=["POST"])
def api_export():
    data = request.json or {}
    token = (data.get("token") or "").strip()
    mode = (data.get("mode") or "catalog").strip().lower()
    sku_list = data.get("skus", [])
    mlb_list = data.get("mlbs", [])
    import_rows = data.get("import_rows") or []
    gumga_token = (data.get("gumga_token") or GUMGA_TOKEN or "").strip()
    any_platform = (data.get("any_platform") or ANYMARKET_PLATFORM or "SELETA").strip()
    reviews = data.get("reviews") or {}
    filter_decision = (data.get("filter_decision") or "all").strip().lower()

    if not token:
        return jsonify({"success": False, "error": "Token do Mercado Livre é obrigatório."}), 400

    # Modo Catálogo (por SKU)
    if mode == "catalog" or sku_list:
        if import_rows:
            sku_from_import = [
                str(r.get("sku") or r.get("id_sku") or "").strip()
                for r in import_rows
                if isinstance(r, dict) and (r.get("sku") or r.get("id_sku"))
            ]
            if sku_from_import:
                sku_list = sku_from_import
        elif not sku_list and mlb_list:
            # Caso o usuário tenha colado SKUs no campo geral
            sku_list = mlb_list

        if not sku_list:
            return jsonify({"success": False, "error": "Informe uma lista de SKUs."}), 400

        audit_items = data.get("audit_items")
        if isinstance(audit_items, list) and audit_items:
            print(f"[EXPORT CATÁLOGO] {len(audit_items)} itens da auditoria | filtro={filter_decision}")
        else:
            print(f"[EXPORT CATÁLOGO] {len(sku_list)} SKUs | filtro={filter_decision}")
        try:
            rows, errors = process_skus_for_catalog_excel(
                sku_list,
                token,
                reviews=reviews or None,
                filter_decision=filter_decision,
                audit_items=audit_items if isinstance(audit_items, list) and audit_items else None,
            )
        except Exception as exc:
            return jsonify({"success": False, "error": f"Erro interno: {str(exc)}"}), 500

        product_count = len(rows) - 1
        if product_count == 0:
            return jsonify({"success": False, "error": "Nenhum produto encontrado para exportar."}), 422

        buf = _write_excel(rows)
        prefix = "Aprovados" if filter_decision == "approved" else ("Reprovados" if filter_decision == "rejected" else "Catalogo_vs_Tradicional")
        filename = f"Relatorio_ML_{prefix}_{uuid.uuid4().hex[:6]}.xlsx"

        response = send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["X-Export-Count"] = str(product_count)
        return response

    # Modo Tradicional (por MLB)
    if import_rows:
        mlb_from_import = [
            str(r.get("mlb") or "").strip().upper()
            for r in import_rows
            if isinstance(r, dict) and r.get("mlb")
        ]
        if mlb_from_import:
            mlb_list = mlb_from_import

    if filter_decision == "approved":
        mlb_list = [m for m in mlb_list if str(reviews.get(m) or "").upper() == "APROVADO"]
        prefix = "Aprovados"
    elif filter_decision == "rejected":
        mlb_list = [m for m in mlb_list if str(reviews.get(m) or "").upper() == "REPROVADO"]
        prefix = "Reprovados"
    elif filter_decision == "pending":
        mlb_list = [m for m in mlb_list if str(reviews.get(m) or "").upper() not in ("APROVADO", "REPROVADO")]
        prefix = "Pendentes"
    else:
        prefix = "Auditoria" if reviews else "MLB"

    if not mlb_list:
        msg = "Nenhum anúncio encontrado para exportar"
        if filter_decision == "approved":
            msg = "Nenhum anúncio foi marcado como Aprovado ainda."
        elif filter_decision == "rejected":
            msg = "Nenhum anúncio foi marcado como Reprovado."
        return jsonify({"success": False, "error": msg}), 400

    print(
        f"[EXPORT] {len(mlb_list)} MLBs (filtro: {filter_decision}) | import={'sim' if import_rows else 'nao'} "
        f"| any={'sim' if gumga_token else 'nao'}"
    )

    try:
        rows, errors = process_mlbs(
            mlb_list,
            token,
            gumga_token=gumga_token or None,
            any_platform=any_platform,
            import_rows=import_rows or None,
            reviews=reviews or None,
        )
    except Exception as exc:
        return jsonify({"success": False, "error": f"Erro interno: {str(exc)}"}), 500

    product_count = len(rows) - 1

    if product_count == 0:
        return jsonify({
            "success": False,
            "error": "Nenhum produto válido encontrado.",
            "warnings": errors,
        }), 422

    if gumga_token and product_count > 0 and errors and all(e.startswith("[ANYMARKET]") for e in errors):
        return jsonify({
            "success": False,
            "error": errors[0],
            "warnings": errors,
        }), 401

    buf = _write_excel(rows)
    filename = f"Relatorio_MLB_{prefix}_{uuid.uuid4().hex[:6]}.xlsx"

    response = send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["X-Export-Count"] = str(product_count)
    response.headers["X-Export-Warnings"] = json.dumps(errors, ensure_ascii=False)
    response.headers["X-Export-AnyMarket"] = "1" if gumga_token else "0"
    response.headers["X-Export-Import"] = "1" if import_rows else "0"
    return response


@app.route("/api/audit", methods=["POST"])
@app.route("/auditarcatalogo/api/audit", methods=["POST"])
def api_audit():
    data = request.json or {}
    token = (data.get("token") or "").strip()
    mode = (data.get("mode") or "catalog").strip().lower()
    sku_list = data.get("skus", [])
    mlb_list = data.get("mlbs", [])
    import_rows = data.get("import_rows") or []
    gumga_token = (data.get("gumga_token") or GUMGA_TOKEN or "").strip()
    any_platform = (data.get("any_platform") or ANYMARKET_PLATFORM or "SELETA").strip()

    if not token:
        return jsonify({"success": False, "error": "Token do Mercado Livre é obrigatório."}), 400

    # Modo 1: Catálogo vs Tradicional ML (por SKU)
    if mode == "catalog" or sku_list:
        if import_rows:
            sku_from_import = [
                str(r.get("sku") or r.get("id_sku") or "").strip()
                for r in import_rows
                if isinstance(r, dict) and (r.get("sku") or r.get("id_sku"))
            ]
            if sku_from_import:
                sku_list = sku_from_import
        elif not sku_list and mlb_list:
            sku_list = mlb_list

        if not sku_list:
            return jsonify({"success": False, "error": "Informe uma lista de SKUs."}), 400

        print(f"[AUDIT CATÁLOGO] {len(sku_list)} SKUs")
        try:
            result = process_skus_for_catalog_audit(
                sku_list,
                token,
                gumga_token=gumga_token,
                any_platform=any_platform,
            )
            return jsonify({
                "success": True,
                "mode": "catalog",
                "summary": result.get("summary", {}),
                "items": result.get("items", []),
                "warnings": result.get("errors", []),
            })
        except Exception as exc:
            return jsonify({"success": False, "error": f"Erro interno: {str(exc)}"}), 500

    # Modo 2: ML vs AnyMarket (por MLB)
    if import_rows:
        mlb_from_import = [
            str(r.get("mlb") or "").strip().upper()
            for r in import_rows
            if isinstance(r, dict) and r.get("mlb")
        ]
        if mlb_from_import:
            mlb_list = mlb_from_import
    elif not mlb_list:
        return jsonify({"success": False, "error": "Informe MLBs ou importe uma planilha."}), 400

    print(
        f"[AUDIT MLB] {len(mlb_list)} MLBs | import={'sim' if import_rows else 'nao'} "
        f"| any={'sim' if gumga_token else 'nao'}"
    )

    try:
        result = process_mlbs_for_audit(
            mlb_list,
            token,
            gumga_token=gumga_token or None,
            any_platform=any_platform,
            import_rows=import_rows or None,
        )
    except Exception as exc:
        return jsonify({"success": False, "error": f"Erro interno: {str(exc)}"}), 500

    items = result.get("items") or []
    errors = result.get("errors") or []

    if gumga_token and errors and all(e.startswith("[ANYMARKET]") for e in errors) and not items:
        return jsonify({
            "success": False,
            "error": errors[0],
            "warnings": errors,
        }), 401

    return jsonify({
        "success": True,
        "mode": "mlb",
        "summary": result.get("summary", {}),
        "items": items,
        "warnings": errors,
    })


@app.route("/api/sync_google_sheet", methods=["POST"])
@app.route("/auditarcatalogo/api/sync_google_sheet", methods=["POST"])
def sync_google_sheet():
    """
    Atualiza o status na planilha do Google (Coluna G = 'Correto')
    ao aprovar um anúncio de catálogo pelo MLB.
    """
    import requests as req
    data = request.get_json() or {}
    webhook_url = data.get("webhook_url", "").strip() or os.environ.get("GOOGLE_SHEET_WEBHOOK_URL", "").strip()
    
    if not webhook_url:
        return jsonify({
            "success": False,
            "error": "URL do Webhook do Google Apps Script não configurada."
        }), 400

    items = data.get("items")
    if not items:
        mlb = data.get("mlb", "").strip()
        sku = data.get("sku", "").strip()
        status = data.get("status", "Correto").strip()
        if not mlb:
            return jsonify({"success": False, "error": "MLB não informado."}), 400
        items = [{"mlb": mlb, "sku": sku, "status": status}]

    try:
        r = req.post(webhook_url, json={"items": items}, timeout=15)
        if r.status_code == 200:
            try:
                res_json = r.json()
            except Exception:
                res_json = {"success": True, "raw": r.text}
            return jsonify({"success": True, "result": res_json})
        else:
            return jsonify({"success": False, "error": f"Erro {r.status_code} ao contatar Google Apps Script: {r.text[:200]}"}), 502
    except Exception as exc:
        return jsonify({"success": False, "error": f"Falha na requisição para a Planilha: {str(exc)}"}), 500


_AI_USER_ERROR = (
    "Falha na pré-validação por IA. Tente novamente ou prossiga com revisão manual."
)


def _validate_one_ad(raw_item: dict) -> dict:
    sku = str((raw_item or {}).get("sku") or "")
    try:
        report = run_prevalidation(raw_item or {})
        sku = str(report.get("sku") or sku)
        return {
            "success": True,
            "sku": sku,
            "report": report,
            "ai_status": "ok",
            "prompt_version": PROMPT_VERSION,
        }
    except PrevalidationError as exc:
        try:
            sku = sku or pair_from_audit_item(raw_item or {}).get("sku") or ""
        except Exception:
            pass
        return {
            "success": False,
            "sku": sku,
            "ai_status": "error",
            "error": _AI_USER_ERROR,
            "detail": exc.detail,
            "prompt_version": PROMPT_VERSION,
        }
    except Exception as exc:
        print(f"[VALIDATE-ADS] erro inesperado: {type(exc).__name__}: {exc}")
        return {
            "success": False,
            "sku": sku,
            "ai_status": "error",
            "error": _AI_USER_ERROR,
            "detail": "upstream_4xx",
            "prompt_version": PROMPT_VERSION,
        }


def _fail_closed_status(detail: str) -> int:
    if detail in ("invalid_payload", "invalid_json"):
        return 422
    if detail == "missing_key":
        return 503
    return 502


@app.route("/api/validate-ads", methods=["POST"])
@app.route("/auditarcatalogo/api/validate-ads", methods=["POST"])
def api_validate_ads():
    if not AI_PREVALIDATION_ENABLED:
        return jsonify({
            "success": False,
            "ai_status": "disabled",
            "error": "Pré-validação IA temporariamente desabilitada.",
            "detail": "ai_disabled",
            "prompt_version": PROMPT_VERSION,
        }), 503

    data = request.json or {}
    raw_items = data.get("items")
    if raw_items is None and (data.get("sku") or data.get("catalogo") or data.get("ml")):
        raw_items = [data]
    if not isinstance(raw_items, list) or not raw_items:
        return jsonify({
            "success": False,
            "ai_status": "error",
            "error": "Informe ao menos um anúncio para pré-validar.",
            "detail": "invalid_payload",
        }), 400

    results = [_validate_one_ad(item if isinstance(item, dict) else {}) for item in raw_items]

    if len(results) == 1:
        item = results[0]
        if item.get("success"):
            return jsonify(item)
        return jsonify(item), _fail_closed_status(str(item.get("detail") or "upstream_4xx"))

    return jsonify({
        "success": True,
        "ai_status": "ok",
        "prompt_version": PROMPT_VERSION,
        "results": results,
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3002))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    print(f"Servidor em http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", debug=debug, port=port)

